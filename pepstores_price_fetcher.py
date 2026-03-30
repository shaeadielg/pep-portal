"""
pepstores_price_fetcher.py
--------------------------
Logs into the Pepstores Centric portal, fetches all SENT POs and their
supplier cost prices in two API calls, then writes an updated .xlsx
with a new "Supplier Cost Price (ZAR)" column.

Usage:
    1. Set USERNAME and PASSWORD below.
    2. Place your exported spreadsheet next to this script (or update INPUT_FILE).
    3. Run: python pepstores_price_fetcher.py
    4. Output: pepstores_prices_updated.xlsx
"""

import requests
import json
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import sys
import os

from dotenv import load_dotenv
import os

load_dotenv()

username = os.getenv("DEFAULT_USERNAME")
password = os.getenv("DEFAULT_PASSWORD")

# --- CONFIGURATION ------------------------------------------------------------
USERNAME   = username
PASSWORD   = password
INPUT_FILE = "Sent_POs.xlsx"

BASE_URL   = "https://pepstores-prod.centricsoftware.com"
API_URL    = f"{BASE_URL}/csi-requesthandler/RequestHandler"
THROTTLE   = 0.3
# ------------------------------------------------------------------------------


def api_post(session, payload):
    """POST to the RequestHandler and return parsed JSON."""
    params  = {"request.preventCache": str(int(time.time() * 1000))}
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Referer":      f"{BASE_URL}/WebAccess/home.html",
    }
    resp = session.post(API_URL, params=params, data=payload, headers=headers, timeout=60)
    resp.raise_for_status()
    text = resp.text.strip().lstrip("(").rstrip(")")
    return json.loads(text)


def login(session):
    """Authenticate and store session cookies."""
    try:
        data = api_post(session, {
            "Fmt.Version": "2",
            "LoginID":     USERNAME,
            "Password":    PASSWORD,
            "Module":      "DataSource",
            "Operation":   "SimpleLogin",
            "OutputJSON":  "1",
        })
        if data.get("Status") == "Successful":
            print("[+] Login successful")
            return True
        print(f"[-] Login failed: {data.get('Status')} - {data.get('Message', '')}")
        return False
    except Exception as e:
        print(f"[-] Login error: {e}")
        return False


def fetch_all_sent_pos(session):
    """
    Fetch all SENT POs in one QueryByXML call.
    Returns list of ResultNode dicts containing PO number and internal URL.
    """
    qry_xml = (
        '<?xml version="1.0" encoding="utf-8" ?>'
        '<Query>'
        '<Node Parameter="Type" Op="EQ" Value="PurchasedOrder"/>'
        '<Attribute Id="State" Op="NE" SValue="PurchasedOrderState:Abandoned"/>'
        '<Attribute Id="po_display_status" Op="EQ" SValue="SENT"/>'
        '</Query>'
    )
    try:
        data = api_post(session, {
            "Fmt.Version":  "2",
            "Fmt.AC.Rights":"Current",
            "Fmt.Attr.Info":"Mid",
            "Module":       "Search",
            "Operation":    "QueryByXML",
            "OutputJSON":   "1",
            "Qry.XML":      qry_xml,
        })
        nodes = data.get("NODES", {}).get("ResultNode", [])
        print(f"[+] Found {len(nodes)} SENT POs on portal")
        return nodes
    except Exception as e:
        print(f"[-] Error fetching PO list: {e}")
        return []


def fetch_prices_batch(session, centric_urls):
    """
    Fetch prices for all internal Centric URLs, batched 20 at a time.
    Returns dict of { centric_url: price_float }
    """
    BATCH_SIZE = 20
    price_map  = {}

    for i in range(0, len(centric_urls), BATCH_SIZE):
        batch = centric_urls[i:i + BATCH_SIZE]
        print(f"  Batch {i//BATCH_SIZE + 1}/{-(-len(centric_urls)//BATCH_SIZE)} ({len(batch)} POs)...")

        payload_list = [
            ("Fmt.Version",   "2"),
            ("Fmt.AC.Rights", "Current"),
            ("Fmt.Attr.Info", "Mid"),
            ("Module",        "Search"),
            ("Operation",     "QueryByURL"),
            ("OutputJSON",    "1"),
        ]
        for url in batch:
            payload_list.append(("Qry.URL", url))

        try:
            data  = api_post(session, payload_list)
            nodes = data.get("NODES", {}).get("ResultNode", [])
            for node in nodes:
                url   = node.get("p_po_url") or node.get("$URL", "")
                price = node.get("p_po_local_avg_cost_price")
                if price is None:
                    price = node.get("p_po_latest_becp")
                if url and not url.startswith("centric://"):
                    price_map[url] = float(price) if price is not None else None
        except Exception as e:
            print(f"  [!] Batch fetch error: {e}")

        time.sleep(THROTTLE)

    return price_map


def main():
    if not os.path.exists(INPUT_FILE):
        print(f"[-] Input file not found: {INPUT_FILE}")
        sys.exit(1)

    # Load spreadsheet
    print(f"[+] Reading spreadsheet: {INPUT_FILE}")
    df = pd.read_excel(INPUT_FILE, dtype={"Supplier PO": str})
    df["Supplier PO"] = df["Supplier PO"].str.strip()

    # Login
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36"
    })

    if not login(session):
        sys.exit(1)

    # Fetch all SENT POs - build PO number to internal URL map
    po_nodes  = fetch_all_sent_pos(session)
    po_to_url = {}
    for node in po_nodes:
        po_num = node.get("$Name") or node.get("ControlNumber") or node.get("p_purchasedorder_mms_no")
        url    = node.get("p_po_url") or node.get("$URL")
        if po_num and url and not url.startswith("centric://"):
            po_to_url[str(po_num).strip()] = url

    print(f"[+] Resolved {len(po_to_url)} PO number to URL mappings")

    if not po_to_url:
        print("[-] No PO URLs resolved - cannot continue.")
        # Debug: show what the first node looked like
        if po_nodes:
            keys = list(po_nodes[0].keys())[:15]
            print(f"    First node keys: {keys}")
        sys.exit(1)

    # Batch fetch prices
    print(f"[+] Fetching prices in batches...")
    url_to_price = fetch_prices_batch(session, list(po_to_url.values()))

    # Build final PO number -> price map
    po_to_price = {po: url_to_price.get(url) for po, url in po_to_url.items()}
    found     = sum(1 for v in po_to_price.values() if v is not None)
    not_found = sum(1 for v in po_to_price.values() if v is None)
    print(f"[+] Prices: {found} found, {not_found} not found")

    # Write results back to spreadsheet
    output_file = "pepstores_prices_updated.xlsx"
    print(f"[+] Writing results to {output_file}")

    def lookup_price(po_val):
        if pd.isna(po_val):
            return None
        try:
            key = str(int(float(str(po_val).strip())))
        except (ValueError, TypeError):
            return None
        return po_to_price.get(key)

    df["Supplier Cost Price (ZAR)"] = df["Supplier PO"].apply(lookup_price)
    df.to_excel(output_file, index=False)

    # Style the new column
    wb = load_workbook(output_file)
    ws = wb.active

    price_col_idx = None
    for col in ws.iter_cols(1, ws.max_column, 1, 1):
        for cell in col:
            if cell.value == "Supplier Cost Price (ZAR)":
                price_col_idx = cell.column
                break

    if price_col_idx:
        header_cell = ws.cell(row=1, column=price_col_idx)
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill("solid", start_color="1F5C99")
        ws.column_dimensions[header_cell.column_letter].width = 28
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=price_col_idx)
            if cell.value is not None:
                cell.number_format = 'R#,##0.0000'

    wb.save(output_file)

    print(f"\n-- Summary -----------------------------------------")
    print(f"  Portal POs found    : {len(po_to_url)}")
    print(f"  Prices found        : {found}")
    print(f"  Not found           : {not_found}")
    print(f"  Output file         : {output_file}")
    print(f"----------------------------------------------------")


if __name__ == "__main__":
    main()
