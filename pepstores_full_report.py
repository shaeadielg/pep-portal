"""
pepstores_full_report.py
------------------------
Fetches all SENT POs from the Pepstores Centric portal and generates
a full breakdown report with supplier, style, qty, and price.

Usage:
    python pepstores_full_report.py
Output:
    pepstores_full_report.xlsx
"""

import requests
import json
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

from dotenv import load_dotenv
import os

load_dotenv()

username = os.getenv("DEFAULT_USERNAME")
password = os.getenv("DEFAULT_PASSWORD")

# --- CONFIGURATION ------------------------------------------------------------
USERNAME   = username
PASSWORD   = password
OUTPUT_FILE = "pepstores_full_report.xlsx"

BASE_URL   = "https://pepstores-prod.centricsoftware.com"
API_URL    = f"{BASE_URL}/csi-requesthandler/RequestHandler"
THROTTLE   = 0.3
BATCH_SIZE = 20
# ------------------------------------------------------------------------------


def api_post(session, payload):
    params  = {"request.preventCache": str(int(time.time() * 1000))}
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Referer":      f"{BASE_URL}/WebAccess/home.html",
    }
    resp = session.post(API_URL, params=params, data=payload, headers=headers, timeout=60)
    resp.raise_for_status()
    text = resp.text.strip().lstrip("(").rstrip(")")
    return json.loads(text)


def login(session):
    try:
        data = api_post(session, {
            "Fmt.Version": "2",
            "LoginID":     USERNAME,
            "Password":    PASSWORD,
            "Module":      "DataSource",
            "Operation":   "SimpleLogin",
            "OutputJSON":  "1",
        })
        if data.get("Status") == "Successful":
            print("[+] Login successful")
            return True
        print(f"[-] Login failed: {data.get('Status')} - {data.get('Message', '')}")
        return False
    except Exception as e:
        print(f"[-] Login error: {e}")
        return False


def fetch_all_sent_pos(session):
    qry_xml = (
        '<?xml version="1.0" encoding="utf-8" ?>'
        '<Query>'
        '<Node Parameter="Type" Op="EQ" Value="PurchasedOrder"/>'
        '<Attribute Id="State" Op="NE" SValue="PurchasedOrderState:Abandoned"/>'
        '<Attribute Id="po_display_status" Op="EQ" SValue="SENT"/>'
        '</Query>'
    )
    try:
        data = api_post(session, {
            "Fmt.Version":  "2",
            "Fmt.AC.Rights":"Current",
            "Fmt.Attr.Info":"Mid",
            "Module":       "Search",
            "Operation":    "QueryByXML",
            "OutputJSON":   "1",
            "Qry.XML":      qry_xml,
        })
        nodes = data.get("NODES", {}).get("ResultNode", [])
        print(f"[+] Found {len(nodes)} SENT POs on portal")
        return nodes
    except Exception as e:
        print(f"[-] Error fetching PO list: {e}")
        return []


def fetch_prices_batch(session, centric_urls):
    price_map = {}
    total_batches = -(-len(centric_urls) // BATCH_SIZE)

    for i in range(0, len(centric_urls), BATCH_SIZE):
        batch = centric_urls[i:i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        print(f"  Fetching batch {batch_num}/{total_batches} ({len(batch)} POs)...")

        payload_list = [
            ("Fmt.Version",   "2"),
            ("Fmt.AC.Rights", "Current"),
            ("Fmt.Attr.Info", "Mid"),
            ("Module",        "Search"),
            ("Operation",     "QueryByURL"),
            ("OutputJSON",    "1"),
        ]
        for url in batch:
            payload_list.append(("Qry.URL", url))

        try:
            data  = api_post(session, payload_list)
            nodes = data.get("NODES", {}).get("ResultNode", [])
            for node in nodes:
                url   = node.get("p_po_url") or node.get("$URL", "")
                price = node.get("p_po_local_avg_cost_price")
                if price is None:
                    price = node.get("p_po_latest_becp")
                if url and not url.startswith("centric://"):
                    price_map[url] = {
                        "price":      float(price) if price is not None else None,
                        "style_code": node.get("p_purchasedorder_style_code", ""),
                        "qty":        node.get("p_po_order_latest_Qty"),
                        "currency":   node.get("p_purchasedorder_lc_currency_lookup", "ZAR"),
                        "status":     node.get("po_display_status", ""),
                        "buyer":      node.get("p_purchasedorder_username", ""),
                        "dc":         node.get("p_purchasedorder_branch_lookup", ""),
                        "season":     node.get("p_purchasedorder_mms_season", ""),
                        "sent_date":  node.get("p_po_sent_date"),
                        "ship_from":  node.get("P_PO_ShipFromDate"),
                        "ship_to":    node.get("P_PO_ShipToDate"),
                        "total_cost": node.get("p_po_supplier_total_fob_cost"),
                        "rsp":        node.get("p_po_rsp"),
                    }
        except Exception as e:
            print(f"  [!] Batch error: {e}")

        time.sleep(THROTTLE)

    return price_map


def ms_timestamp_to_date(ms):
    """Convert Centric millisecond timestamp to readable date string."""
    if not ms or ms == 0:
        return ""
    try:
        import datetime
        dt = datetime.datetime.utcfromtimestamp(int(ms) / 1000)
        return dt.strftime("%d/%m/%Y")
    except:
        return ""


def style_report(ws, row_count):
    """Apply formatting to the report worksheet."""
    # Header row styling
    header_fill = PatternFill("solid", start_color="1F5C99")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Alternating row colours
    fill_light = PatternFill("solid", start_color="F2F7FC")
    fill_white = PatternFill("solid", start_color="FFFFFF")
    data_font  = Font(size=10)

    for row in range(2, row_count + 2):
        fill = fill_light if row % 2 == 0 else fill_white
        for cell in ws[row]:
            cell.fill = fill
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Column widths
    col_widths = {
        "A": 12,   # PO Number
        "B": 14,   # Style Code
        "C": 45,   # Style Description
        "D": 10,   # Season
        "E": 12,   # Qty
        "F": 16,   # Unit Price (ZAR)
        "G": 16,   # Total Cost (ZAR)
        "H": 10,   # RSP
        "I": 12,   # Currency
        "J": 20,   # PEP Buyer
        "K": 20,   # DC
        "L": 14,   # Ship From
        "M": 14,   # Ship To
        "N": 14,   # PEP Sent Date
        "O": 10,   # Status
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Number formatting for price columns
    for row in range(2, row_count + 2):
        ws[f"F{row}"].number_format = 'R#,##0.0000'
        ws[f"G{row}"].number_format = 'R#,##0.00'
        ws[f"H{row}"].number_format = 'R#,##0.00'
        ws[f"E{row}"].number_format = '#,##0'


def main():
    # Login
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36"
    })

    if not login(session):
        sys.exit(1)

    # Fetch all SENT POs
    po_nodes = fetch_all_sent_pos(session)
    if not po_nodes:
        print("[-] No POs returned.")
        sys.exit(1)

    # Build PO number -> internal URL map
    po_to_url = {}
    po_meta   = {}   # internal url -> basic info from list response
    for node in po_nodes:
        po_num = node.get("$Name") or node.get("ControlNumber")
        url    = node.get("p_po_url") or node.get("$URL")
        if po_num and url and not url.startswith("centric://"):
            po_to_url[str(po_num).strip()] = url
            po_meta[url] = {
                "po_number": str(po_num).strip(),
                "style":     node.get("p_purchasedorder_style_code", ""),
            }

    print(f"[+] Resolved {len(po_to_url)} PO URLs")

    # Batch fetch full details + prices
    print(f"[+] Fetching full PO details in batches...")
    url_to_detail = fetch_prices_batch(session, list(po_to_url.values()))

    # Build report rows
    print(f"[+] Building report...")
    rows = []
    for po_num, url in sorted(po_to_url.items(), key=lambda x: x[0]):
        detail = url_to_detail.get(url, {})
        rows.append({
            "PO Number":        po_num,
            "Style Code":       detail.get("style_code") or po_meta.get(url, {}).get("style", ""),
            "Season":           detail.get("season", ""),
            "Qty":              int(detail["qty"]) if detail.get("qty") else "",
            "Unit Price (ZAR)": detail.get("price"),
            "Total Cost (ZAR)": detail.get("total_cost"),
            "RSP":              detail.get("rsp"),
            "Currency":         detail.get("currency", "ZAR"),
            "PEP Buyer":        detail.get("buyer", ""),
            "DC":               detail.get("dc", ""),
            "Ship From":        ms_timestamp_to_date(detail.get("ship_from")),
            "Ship To":          ms_timestamp_to_date(detail.get("ship_to")),
            "PEP Sent Date":    ms_timestamp_to_date(detail.get("sent_date")),
            "Status":           detail.get("status", "SENT"),
        })

    df = pd.DataFrame(rows, columns=[
        "PO Number", "Style Code", "Season", "Qty",
        "Unit Price (ZAR)", "Total Cost (ZAR)", "RSP", "Currency",
        "PEP Buyer", "DC", "Ship From", "Ship To", "PEP Sent Date", "Status"
    ])

    # Write to Excel
    print(f"[+] Writing {len(df)} rows to {OUTPUT_FILE}...")
    df.to_excel(OUTPUT_FILE, index=False, sheet_name="SENT POs")

    # Apply styling
    wb = load_workbook(OUTPUT_FILE)
    ws = wb["SENT POs"]
    style_report(ws, len(df))

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Report Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)

    summary_data = [
        ("Total SENT POs",        len(df)),
        ("Prices Found",          df["Unit Price (ZAR)"].notna().sum()),
        ("Prices Missing",        df["Unit Price (ZAR)"].isna().sum()),
        ("Total Qty",             int(df["Qty"].replace("", 0).fillna(0).sum())),
        ("Total Cost Value (ZAR)",df["Total Cost (ZAR)"].sum()),
        ("Unique Style Codes",    df["Style Code"].nunique()),
        ("Unique Buyers",         df["PEP Buyer"].nunique()),
    ]

    for i, (label, value) in enumerate(summary_data, start=3):
        ws_sum[f"A{i}"] = label
        ws_sum[f"B{i}"] = value
        ws_sum[f"A{i}"].font = Font(bold=True, size=10)
        if "Cost" in label:
            ws_sum[f"B{i}"].number_format = 'R#,##0.00'
        elif "Qty" in label:
            ws_sum[f"B{i}"].number_format = '#,##0'

    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 20

    wb.save(OUTPUT_FILE)

    print(f"\n-- Summary -----------------------------------------")
    print(f"  Total SENT POs      : {len(df)}")
    print(f"  Prices found        : {df['Unit Price (ZAR)'].notna().sum()}")
    print(f"  Prices missing      : {df['Unit Price (ZAR)'].isna().sum()}")
    print(f"  Total Qty           : {int(df['Qty'].replace('', 0).fillna(0).sum()):,}")
    print(f"  Output file         : {OUTPUT_FILE}")
    print(f"----------------------------------------------------")


if __name__ == "__main__":
    main()
