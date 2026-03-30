"""
pepstores_sku_report.py
-----------------------
Joshtex - PEP Stores SKU Detail Report
Northern Textile Mills SA (PTY) Ltd

Fetches all SENT POs and their SKU-level order line details
(MMS SKU#, description, size, colour, qty, carton qty, cost, RSP etc.)

Output: pepstores_sku_report.xlsx
"""

import requests
import json
import time
import datetime
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv
import os

load_dotenv()

username = os.getenv("DEFAULT_USERNAME")
password = os.getenv("DEFAULT_PASSWORD")

# ── Credentials ────────────────────────────────────────────────────────────────
try:
    import sqlite3
    conn = sqlite3.connect(r"P:\permissions.db")
    cur  = conn.cursor()
    cur.execute("SELECT username, password FROM pep_import WHERE app_name='Pep Import' LIMIT 1")
    row = cur.fetchone()
    conn.close()
    USERNAME, PASSWORD = (row[0], row[1]) if row else (username, password)
except Exception:
    USERNAME, PASSWORD = username, password

# ── Config ──────────────────────────────────────────────────────────────────────
BASE_URL   = "https://pepstores-prod.centricsoftware.com"
API_URL    = f"{BASE_URL}/csi-requesthandler/RequestHandler"
THROTTLE   = 0.3
BATCH_SIZE = 20
OUTPUT     = "pepstores_sku_report.xlsx"

# ── API helpers ─────────────────────────────────────────────────────────────────
def api_post(session, payload):
    params  = {"request.preventCache": str(int(time.time() * 1000))}
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Referer":      f"{BASE_URL}/WebAccess/home.html",
    }
    resp = session.post(API_URL, params=params, data=payload, headers=headers, timeout=60)
    resp.raise_for_status()
    text = resp.text.strip().lstrip("(").rstrip(")")
    return json.loads(text)


def login(session):
    data = api_post(session, {
        "Fmt.Version": "2", "LoginID": USERNAME, "Password": PASSWORD,
        "Module": "DataSource", "Operation": "SimpleLogin", "OutputJSON": "1",
    })
    if data.get("Status") == "Successful":
        print("[+] Login successful")
        return True
    print(f"[-] Login failed: {data.get('Status')}")
    return False


def fetch_sent_pos(session):
    qry_xml = (
        '<?xml version="1.0" encoding="utf-8" ?><Query>'
        '<Node Parameter="Type" Op="EQ" Value="PurchasedOrder"/>'
        '<Attribute Id="State" Op="NE" SValue="PurchasedOrderState:Abandoned"/>'
        '<Attribute Id="po_display_status" Op="EQ" SValue="SENT"/>'
        '</Query>'
    )
    data = api_post(session, {
        "Fmt.Version": "2", "Fmt.AC.Rights": "Current", "Fmt.Attr.Info": "Mid",
        "Module": "Search", "Operation": "QueryByXML", "OutputJSON": "1",
        "Qry.XML": qry_xml,
    })
    nodes = data.get("NODES", {}).get("ResultNode", [])
    print(f"[+] Found {len(nodes)} SENT POs")
    return nodes


def fetch_order_lines_batch(session, po_products_urls, po_meta):
    """
    Fetch Order child nodes for a batch of POProduct URLs.
    Returns list of row dicts.
    """
    rows = []
    total = -(-len(po_products_urls) // BATCH_SIZE)

    for i in range(0, len(po_products_urls), BATCH_SIZE):
        batch = po_products_urls[i:i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        print(f"  Order lines batch {batch_num}/{total}...")

        payload_list = [
            ("Fmt.Version", "2"),
            ("Fmt.AC.Rights", "Current"),
            ("Fmt.Attr.Info", "Mid"),
            ("Module", "Search"),
            ("Operation", "QueryByXML"),
            ("OutputJSON", "1"),
            ("Fmt.Complete", "Ref"),
            ("Fmt.Complete.Max", "3999"),
            ("Fmt.Crew", "Name"),
            ("Crew.Scope", "Result"),
            ("Qry.Limit.Filter", "10000"),
            ("Qry.Limit.Begin", "1"),
            ("Qry.Limit.End", "2000"),
        ]

        for po_prod_url in batch:
            payload_list.append(("Qry.Limit.Path", f"{po_prod_url}?Path=Child%3AOrders"))

        # Dep.Path to pull in related nodes (size, color, style)
        payload_list += [
            ("Dep.Path", "Child:RealizedProduct/Child:RealizedSize"),
            ("Dep.Path", "Child:RealizedProduct/Child:RealizedColor"),
            ("Dep.Path", "Child:POColor"),
            ("Dep.Path", "Child:PO"),
        ]

        payload_list.append(("Qry.XML",
            '<?xml version="1.0" encoding="utf-8" ?>'
            '<Query Direct="true">'
            '<OR>'
            '<Node Parameter="Type" Op="EQ" Value="OrderTableSpanner"/>'
            '<AND><Node Parameter="Type" Op="EQ" Value="Order"/></AND>'
            '</OR>'
            '<OrderByAttribute Id="P_SKU_Size_Sort_Order" Path="(Order)Child:RealizedProduct(SKU)"/>'
            '</Query>'
        ))

        try:
            data = api_post(session, payload_list)

            result_nodes = data.get("NODES", {}).get("ResultNode", [])
            related_nodes = data.get("NODES", {}).get("Node", [])

            # Build lookup maps from related nodes
            sku_map      = {}  # url -> SKU node
            size_map     = {}  # url -> size name
            colorway_map = {}  # url -> colorway node
            po_node_map  = {}  # url -> PO node

            for node in related_nodes:
                t = node.get("__DomainKey__", "")
                url = node.get("$URL", "")
                if t == "SKU":
                    sku_map[url] = node
                elif t == "ProductSize":
                    size_map[url] = node.get("$Name", "")
                elif t == "Colorway":
                    colorway_map[url] = node
                elif t == "PurchasedOrder":
                    po_node_map[url] = node

            for order in result_nodes:
                if order.get("__DomainKey__") != "Order":
                    continue

                po_url   = str(order.get("p_po_url", "")).strip()
                po_node  = po_node_map.get(po_url, {})
                po_num   = order.get("P_Order_PurchOrderNumber", "") or po_node.get("$Name", "")

                # SKU details
                sku_url  = str(order.get("RealizedProduct", "")).strip()
                sku_node = sku_map.get(sku_url, {})

                # Size
                size_url  = str(sku_node.get("RealizedSize", "")).strip()
                size_name = size_map.get(size_url, "")
                if not size_name:
                    # Try parsing from SKU node name e.g. "-11865 CC:1 - ENTRY FACECLOTH - GREEN (S26)-30X30CM"
                    sku_node_name = sku_node.get("$Name", "")
                    if "-" in sku_node_name:
                        size_name = sku_node_name.split("-")[-1].strip()

                # Colour
                color_url  = str(sku_node.get("RealizedColor", "")).strip()
                color_node = colorway_map.get(color_url, {})
                colour     = color_node.get("P_CW_COLOUR_FAMILY", "")
                color_spec = color_node.get("ColorSpecification_copy", "")

                rows.append({
                    "PO Number":          po_num,
                    "Style Code":         po_node.get("p_purchasedorder_style_code", "") or sku_node.get("p_sku_style_erp_code", ""),
                    "MMS SKU #":          order.get("P_Order_SkuNumber", ""),
                    "Description":        sku_node.get("P_SKU_Description", "") or order.get("$Name", ""),
                    "Size":               size_name,
                    "Colour":             colour,
                    "Colour Spec":        color_spec,
                    "Latest Qty":         int(order["p_order_latest_qty"]) if order.get("p_order_latest_qty") else "",
                    "Carton Qty":         int(order["p_order_carton_qty"]) if order.get("p_order_carton_qty") else "",
                    "In Pack Qty":        int(order["p_order_inpack_qty"]) if order.get("p_order_inpack_qty") else "",
                    "Unit Cost (ZAR)":    order.get("p_order_latest_becp"),
                    "Total Cost (ZAR)":   order.get("p_order_latest_becp_total"),
                    "RSP (excl)":         order.get("p_order_rsp"),
                    "RSP (incl)":         order.get("p_order_incl_rsp"),
                    "GP %":               round(order["p_order_gp"] * 100, 2) if order.get("p_order_gp") else "",
                    "Season":             po_node.get("p_purchasedorder_mms_season", ""),
                    "Department":         po_node.get("P_PO_CC_DeptNumber", ""),
                    "DC":                 po_node.get("p_purchasedorder_branch_lookup", ""),
                    "Supplier Contact":   po_node.get("p_purchaseorder_supplier_username", ""),
                    "PEP Buyer":          po_node.get("p_purchasedorder_username", ""),
                    "Status":             po_node.get("po_display_status", "SENT"),
                })

        except Exception as e:
            print(f"  [!] Batch error: {e}")

        time.sleep(THROTTLE)

    return rows


def style_sheet(ws, df):
    BLACK  = "1E1E1E"
    GOLD   = "D4AF37"
    WHITE  = "FFFFFF"
    LGREY  = "F5F5F5"

    hdr_fill = PatternFill("solid", fgColor=BLACK)
    hdr_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    hdr_align = Alignment(horizontal="left", vertical="center")

    alt_fill = PatternFill("solid", fgColor=LGREY)
    data_font = Font(name="Calibri", size=9)
    data_align = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="DDDDDD")
    border = Border(bottom=thin)

    num_cols = {"Unit Cost (ZAR)", "Total Cost (ZAR)", "RSP (excl)", "RSP (incl)", "GP %",
                "Latest Qty", "Carton Qty", "In Pack Qty"}
    right_align = Alignment(horizontal="right", vertical="center")

    for cell in ws[1]:
        cell.fill    = hdr_fill
        cell.font    = hdr_font
        cell.alignment = hdr_align

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for cell in row:
            col_name = df.columns[cell.column - 1]
            cell.font      = data_font
            cell.alignment = right_align if col_name in num_cols else data_align
            cell.border    = border
            if fill:
                cell.fill = fill

    # Column widths
    widths = {
        "PO Number": 12, "Style Code": 12, "MMS SKU #": 12,
        "Description": 40, "Size": 12, "Colour": 14, "Colour Spec": 28,
        "Latest Qty": 12, "Carton Qty": 12, "In Pack Qty": 12,
        "Unit Cost (ZAR)": 16, "Total Cost (ZAR)": 16,
        "RSP (excl)": 12, "RSP (incl)": 12, "GP %": 10,
        "Season": 10, "Department": 26, "DC": 10,
        "Supplier Contact": 18, "PEP Buyer": 18, "Status": 10,
    }
    for i, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col_name, 14)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20


def main():
    print(f"[+] PEP Stores SKU Detail Report")
    print(f"[+] Credentials: {USERNAME}")

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    })

    if not login(session):
        sys.exit(1)

    po_nodes = fetch_sent_pos(session)
    if not po_nodes:
        print("[-] No POs returned.")
        sys.exit(1)

    # Build POProduct URL list (one per PO — each PO has one POProduct)
    po_products = []
    for node in po_nodes:
        po_prods = node.get("POProducts", [])
        for pp_url in po_prods:
            po_products.append(pp_url)

    print(f"[+] Fetching order lines for {len(po_products)} POProduct nodes...")
    rows = fetch_order_lines_batch(session, po_products, {})

    if not rows:
        print("[-] No order line data returned.")
        sys.exit(1)

    print(f"[+] Total SKU lines: {len(rows)}")

    df = pd.DataFrame(rows, columns=[
        "PO Number", "Style Code", "MMS SKU #", "Description",
        "Size", "Colour", "Colour Spec",
        "Latest Qty", "Carton Qty", "In Pack Qty",
        "Unit Cost (ZAR)", "Total Cost (ZAR)", "RSP (excl)", "RSP (incl)", "GP %",
        "Season", "Department", "DC",
        "Supplier Contact", "PEP Buyer", "Status",
    ])

    # Sort by PO Number then Size
    df.sort_values(["PO Number", "Size"], inplace=True)

    # Write to Excel
    df.to_excel(OUTPUT, index=False, sheet_name="SKU Detail")

    wb = load_workbook(OUTPUT)
    ws = wb["SKU Detail"]
    style_sheet(ws, df)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Report Generated", datetime.datetime.now().strftime("%d %b %Y %H:%M")])
    ws2.append(["Total SKU Lines", len(df)])
    ws2.append(["Total POs", df["PO Number"].nunique()])
    ws2.append(["Unique Style Codes", df["Style Code"].nunique()])
    ws2.append(["Total Qty", int(df["Latest Qty"].sum()) if df["Latest Qty"].dtype != "object" else ""])
    ws2.append(["Total Cost (ZAR)", round(df["Total Cost (ZAR)"].sum(), 2)])
    for row in ws2.iter_rows():
        for cell in row:
            cell.font = Font(name="Calibri", size=10)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18

    wb.save(OUTPUT)
    print(f"[+] Saved: {OUTPUT}")
    print(f"\n-- Complete ----------------------------------------")
    print(f"  SKU lines    : {len(df)}")
    print(f"  POs          : {df['PO Number'].nunique()}")
    print(f"  Output       : {OUTPUT}")
    print(f"----------------------------------------------------")


if __name__ == "__main__":
    main()
